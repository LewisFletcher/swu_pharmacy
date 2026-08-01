import openpyxl
from django.core.management.base import BaseCommand, CommandError

from pharmacy.medication_import import load_medications_from_workbook


class Command(BaseCommand):
    help = (
        "Load/update Medication records from an Excel file with 'Name' and "
        "'Description' columns, e.g.: "
        "Name = 'Albon 15 gm (Sulfadimethoxine)', "
        "Description = 'Give 25 mg/lb BW orally for a max of 5 days. MILK = 60 HRS, MEAT=7D'"
    )

    def add_arguments(self, parser):
        parser.add_argument("path", help="Path to the .xlsx file")
        parser.add_argument("--sheet", default=None, help="Sheet name (defaults to the active sheet)")
        parser.add_argument(
            "--dry-run", action="store_true",
            help="Preview created/updated counts without saving anything",
        )

    def handle(self, *args, **options):
        path = options["path"]
        try:
            workbook = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {path}")
        except Exception as exc:
            raise CommandError(f"Couldn't open '{path}' as an Excel file: {exc}")

        if options["sheet"] and options["sheet"] not in workbook.sheetnames:
            raise CommandError(
                f"Sheet '{options['sheet']}' not found. Available sheets: {', '.join(workbook.sheetnames)}"
            )

        summary = load_medications_from_workbook(
            workbook, sheet_name=options["sheet"], dry_run=options["dry_run"],
        )

        verb = "Would create/update" if options["dry_run"] else "Created/updated"
        self.stdout.write(self.style.SUCCESS(
            f"{verb}: {summary['created']} new, {summary['updated']} existing."
        ))
        if summary["skipped"]:
            self.stdout.write(self.style.WARNING(
                f"Skipped {len(summary['skipped'])} row(s) with no usable name: "
                f"{summary['skipped'][:10]}{'...' if len(summary['skipped']) > 10 else ''}"
            ))
